#!/usr/bin/env python
# coding: utf-8

# In[107]:

import os
import requests
from bs4 import BeautifulSoup
import pandas as pd
import re
import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from textblob import TextBlob




headers = {'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.75 Safari/537.36'}
url = 'https://steamcommunity.com/app/1094520/positivereviews/?filterLanguage=english&p=1&browsefilter=toprated'


page = requests.get(url)
soup = BeautifulSoup(page.content, 'html5lib') # If this line causes an error, run 'pip install html5lib' or install html5lib 


browser = webdriver.Chrome(executable_path=r'C:\Users\Amrutha\Downloads\chromedriver_win32\chromedriver')

browser.get(url)
time.sleep(1)

elem = browser.find_element_by_tag_name("body")
last_height = browser.execute_script("return document.body.scrollHeight")

while True:
     #Scroll down to bottom
     browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")

     # Wait to load page
     time.sleep(5)

    # Calculate new scroll height and compare with last scroll height
     new_height = browser.execute_script("return document.body.scrollHeight")
     if new_height == last_height:
         break
     last_height = new_height

soup = BeautifulSoup(browser.page_source, 'html5lib')

name_offline = soup.find_all('div',class_='apphub_CardContentAuthorName offline ellipsis')


cust_link = []
cust_name = []
for i in range (0,len(name_offline)):
    cust_name.append(name_offline[i].get_text())
    for a_elm in name_offline[i].select("a"):
        cust_link.append(a_elm.attrs["href"])
        break
    

review = soup.find_all('div',class_ = 'apphub_CardTextContent')
review


cust_review = []
cust_review_date = []
cust_review_type = []
for i in range (0,len(review)):
    if review[i].get_text().split("\n").__len__() >= 4:
        if review[i].get_text().split("\n")[3].__contains__("Early Access Review"):
            cust_review.append(review[i].get_text().split("\n")[4])
            cust_review_date.append(review[i].get_text().split("\n")[1])
            cust_review_type.append(review[i].get_text().split("\n")[2]+","+review[i].get_text().split("\n")[3])
        else:
            cust_review.append(review[i].get_text().split("\n")[3])
            cust_review_date.append(review[i].get_text().split("\n")[1])
            cust_review_type.append(review[i].get_text().split("\n")[2])
    else:
        cust_review.append(review[i].get_text().split("\n")[2])
        cust_review_date.append(review[i].get_text().split("\n")[1])
        cust_review_type.append("NA")
    #print(review[i].get_text().split("\n")[3])
cust_review


#cust_review[:] = [review.lstrip('\n') for review in cust_review]
#cust_review


#cust_review[:] = [review.rstrip('\n\t') for review in cust_review]
#cust_review


# Call a Workbook() function of openpyxl  
# to create a new blank Workbook object 
wb = openpyxl.Workbook() 
  
# Get workbook active sheet   
# from the active attribute 
sheet = wb.active 
  
# Cell objects also have row, column 
# and coordinate attributes that provide 
# location information for the cell. 
  
# Note: The first row or column integer 
# is 1, not 0. Cell object is created by 
# using sheet object's cell() method. 
i = 1
for key in cust_name:
    j = 1
    if i == 1:
        c1 = sheet.cell(row = i, column = j) 
        c1.value = "USERID"
        j+=1
        
        c2 = sheet.cell(row = i , column = j) 
        c2.value = "DATE"
        j+=1
        
        c3 = sheet.cell(row = i , column = j) 
        c3.value = "TYPE"
        j+=1
        
        c4 = sheet.cell(row = i , column = j) 
        c4.value = "REVIEW"
        j+=1
        
        c5 = sheet.cell(row = i, column = j)
        c5.value = "USER LINK"
        j+=1
        
        c6 = sheet.cell(row = i , column = j) 
        c6.value = "REVIEW LENGTH"
        j+=1
        
        c7 = sheet.cell(row = i , column = j)
        c7.value = "LANGUAGE"
        
        i+=1
        continue
        
    c1 = sheet.cell(row = i, column = j) 
    c1.value = key
    j+=1
    
    c2 = sheet.cell(row = i , column = j) 
    c2.value = cust_review_date[i-1]
    j+=1
    
    c3 = sheet.cell(row = i , column = j) 
    c3.value = cust_review_type[i-1]
    j+=1
    
    c4 = sheet.cell(row = i , column = j) 
    c4.value = cust_review[i-1]
    j+=1
    
    c5 = sheet.cell(row = i, column = j)
    c5.value = cust_link[i-1]
    j+=1
    
    c6 = sheet.cell(row = i , column = j) 
    c6.value = str(cust_review[i-1]).__len__()
    j+=1
    
    c7 = sheet.cell(row = i , column = j)
    try:
        if str(cust_review[i-1]).__len__() > 4:
            b = TextBlob(cust_review[i-1])
            c7.value = b.detect_language()
        else:
            c7.value = "NA"
    except Exception as e:
        c7.value = "NA"
        time.sleep(5)
        
    i+=1

path = os.getcwd()
wb.save(path+"\\steamcommunity.xlsx") 
browser.close()

# In[ ]:




